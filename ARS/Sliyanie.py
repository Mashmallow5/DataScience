import pandas as pd
import xlrd
from openpyxl import Workbook
from os import listdir
import re


#wb = xlrd.open_workbook('D:\\YourBoardingPassDotAero\YourBoardingPassDotAero-2017-01-01.xlsx')
#sh = wb.sheet_by_index(0)
Et = ['PaxNameFirst', 'PaxNameSecond', 'SEX', 'DepartDate', 'DepartTime', 'FlightCode',
              'From', 'Dest', 'Code',  'e-Ticket', 'SEAT',
              'TrvCls', 'SEQUENCE']

work = Workbook()
#sheet = work.add_sheet('sh1')
ws = work.active

#for i in range(len(Et)):
    #sheet.write(0, i, Et[i])

ws.append(Et)

#work.save('D:\\out.xls')    
    
path = 'D:\\YourBoardingPassDotAero_copy2/'


#index = 1
for f in listdir(path):
    wb = xlrd.open_workbook(path+f, on_demand = True)
    for sh in wb.sheets():

        buf = sh.cell(2,1).value
        buf_spl = re.split(' ', buf)
        
        if len(buf_spl) == 3:
            for i in buf_spl:
                if len(i) ==1:
                    buf_spl.remove(i)
        Sec_name = buf_spl[0]
        Fs_name = buf_spl[1]
            
        
        row = [Fs_name, Sec_name, sh.cell(2,0).value, sh.cell(8,0).value, sh.cell(8,2).value, sh.cell(4,0).value,
               sh.cell(6,3).value, sh.cell(6,7).value, sh.cell(12,1).value, sh.cell(12,4).value, sh.cell(10,7).value,
               sh.cell(2,7).value, sh.cell(0,7).value]

        #for j in range(len(row)):
            #sheet.write(index, j, row[j])
        
        ws.append(row)
        #index += 1

work.save('D:\\out3.xlsx')
